import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import  Workbook,load_workbook

# Veri dosyamızı cagirdik.

arac = load_workbook("veri.xlsx") 
sheet = arac.active
sheet = arac['Veri Seti 2'] # İslenmis veriyi yazdiracagimiz excel sayfasini sectik.


dataset = pd.read_excel('veri.xlsx') # Veriyi okuyup dataset degiskenimize atadik.
X = dataset.iloc[: ,:-1].values # X ve y degsikenlerine bagimli ve bagimsiz degiskenlerimizi atadik. 
y = dataset.iloc[:, -1].values
# print(X)
# print(y)

# Encoding islemi secili sutunlardaki string verilerimizi sayisal degere dönüstürdük, tekrar degiskenimize atadik.

from sklearn.compose import ColumnTransformer 
from sklearn.preprocessing import OneHotEncoder
ct = ColumnTransformer(transformers=[('encoder', OneHotEncoder(), [0,1])], remainder='passthrough')
X = np.array(ct.fit_transform(X))
# print(X)

# Test ve Train kısımlarımıza verileri 80/20 oraninda böldük.

from sklearn.model_selection import train_test_split 
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size = 0.2, random_state = 1)
# print(X_train)
# print(X_test)
# print(y_train)
# print(y_test)

# Secili sutunlardaki verilerimizi belirli bir araliga kisitladik.

from sklearn.preprocessing import StandardScaler
sc = StandardScaler()
X_train[:, 2:] = sc.fit_transform(X_train[:, 2:])
X_test[:, 2:] = sc.fit_transform(X_test[:, 2:])
print(X_train)
print(X_test)

# İslenmis verilerimizi for dögüleri yardimiyla dosyamıza yazdiriyoruz.

for i in range(1,len(X_train)+1): # Ilk önce X_train dizisindeki verilerimizi yazdiriyoruz.
    for j in range(1,6):
        sheet.cell(row=i, column=j).value = X_train[i-1,j-1]
        
for i in range(1,len(X_test)+1): # Yazdırmaya kaldigi yerden devam ettiriyoruz.       
    sheet.append([X_test[i-1,0],X_test[i-1,1],X_test[i-1,2],X_test[i-1,3],X_test[i-1,4]]) 
    
for i in range(1,len(y_train)+1): # y_train degiskenimizi 6. sutuna yazdirdik.
    sheet.cell(row=i, column=6).value = y_train[i-1]
     
for i,j in zip(range(len(y_train)+1,len(y_test)+len(y_train)+1),range(0,len(y_train))): # y_train'in sonundan,kaldigi yerden dizi uzunlugunca devam ettirip verilerimizi yazdirdik.
    sheet.cell(row=i, column=6).value = y_train[j]                                 
    
        
        
# Dosyamızı kaydedip kapatiyoruz.
arac.save("veri.xlsx")
arac.close()