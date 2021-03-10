import requests
from  bs4 import BeautifulSoup
from openpyxl import  Workbook,load_workbook

# openpyxl kutuphanesini kullanarak verileri yazdıracagmız excel dosyasını olusturuyoruz.

arac = Workbook() 
sheet = arac.active
sheet.title = 'Veri Seti 1' 
sheet = arac.create_sheet('Veri Seti 2') # Islenmis verileri yazdırmak icin ikinci bir sayfa oluşturduk.
sheet = arac["Veri Seti 1"] # Kullanacagımız sayfayı seciyoruz.

# Cekilecek veriler icin dizi oluşturuldu.

modelbilgileri =[]
fiyat = []
yıl_km= []

# for dögüsünü kullanarak sayfa sayfa veri almayı sagladık.
for i in range(1,51):
    url = "https://www.arabam.com/ikinci-el/otomobil/renault-megane?take=50&page={}".format(i) # i degiskeni kullanılarak sayfa linki degistirildi.
    response = requests.get(url)    # Web sitesine istek atıldı ve icerik soup degiskenine atandi.
    content = response.content
    soup = BeautifulSoup(content,'html.parser')

    modelbilgileri.extend(soup.find_all('h3',class_='crop-after')) # Model bilgileri(Marka,Model,Motor Hacmi bilgileri) 'h3' tag'ina ait class bilgisi icerisindeki veriler diziye eklendi.

    fiyat.extend(soup.find_all('span', class_='db no-wrap')) # Fiyat bilgileri diziye eklendi.

    yıl_km.extend(soup.find_all('td',class_='listing-text pl8 pr8 tac pr')) # Yıl ve km bilgileri diziye atandı.

# Karisik bilgileri ayırıp kullacagimiz diziler olusturuldu.

    yıl = [] 
    km = []
    marka = []
    model = []
    motor = []


    for mr in range(0,len(modelbilgileri)): # Karisik diziden marka bilgisi cekildi, diziye eklendi.
        temp1 = modelbilgileri[mr].text
        marka.append(temp1[:7])

    for md in range(0,len(modelbilgileri)): # Karisik diziden model bilgisi cekildi, diziye eklendi.
        temp2 = modelbilgileri[md].text
        model.append(temp2[8:14])

    for mt in range(0,len(modelbilgileri)): # Karisik diziden motor hacmi bilgisi cekildi, diziye eklendi.
        temp3 = modelbilgileri[mt].text
        motor.append(temp3[15:18])

    for j in range(0, len(yıl_km), 4): # Karisik diziden yıl bilgisi cekildi, diziye eklendi.
        #print(yıl_km[j].text)
        yıl.append(yıl_km[j].text)

    for k in range(1, len(yıl_km), 4):# Karisik diziden km bilgisi cekildi, diziye eklendi.
        #print(yıl_km[k].text)
        km.append(yıl_km[k].text)
        

# for dögüsü icerisinde zip metodu ile diziler tek döngüde birlestirildi her turda her diziden gecerli degisken kullanilarak excele bu veriler yazdirildi.
print(len(modelbilgileri))
print(len(yıl_km))
print(len(motor))
print(len(yıl))
for i,mr,md,mt,yl,km,fyt in zip(range(1,len(marka)+1),marka,model,motor,yıl,km,fiyat):
        fyt = fyt.text
        fyt = fyt.replace(" TL","") # Fiyat bilgisindeki string ifade replace() metodu ile cikartildi.
        fyt = fyt.replace(".","")
        fyt = fyt.strip()
        km = km.replace(".","")
        km = km.strip()
        
        sheet.cell(row=i, column=1).value = mr
        sheet.cell(row=i, column=2).value = md
        sheet.cell(row=i, column=3).value = mt
        sheet.cell(row=i, column=4).value = yl
        sheet.cell(row=i, column=5).value = km
        sheet.cell(row=i, column=6).value = fyt

# Excel dosyamizi kaydettik.

arac.save("veri.xlsx")
arac.close()










